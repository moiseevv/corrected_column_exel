from openpyxl import load_workbook as lw
import os


def one_exel(file_exel):
    wb = lw(file_exel)
    ws = wb.active

    print("***********************************")
    print(" Число строк в файле", file_exel, "  -  ", ws.max_row + 1)
    col_inkom = 0

    for i in range(1, ws.max_row + 1):
        if "с/с" in str(ws.cell(i, 4).value):
            element = ws.cell(i, 4).value
            mas_el = element.split(",")
            for el in range(0, len(mas_el)):
                if "с/с" in str(mas_el[el]):
                    mas_el.pop(el)
                    new_value = ",".join(mas_el)
                    ws.cell(i, 4).value = new_value
                    col_inkom += 1
                    break

    new_file = file_exel.replace(".xlsx", "_result.xlsx")
    wb.save(new_file)

    print(" Исправленых строк - ", col_inkom)


for l in os.listdir():
    if ".xlsx" in l:
        one_exel(l)
